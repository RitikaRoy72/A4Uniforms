from openpyxl import load_workbook
from models import db, UniformType, SupplyInventory


def import_supply_sheet(file):
    wb = load_workbook(file)
    sheet = wb.active

    sizes = parse_size_headers(sheet)
    parse_inventory(sheet, sizes)

    db.session.commit()


def parse_size_headers(sheet):
    sizes = []
    for col in range(2, sheet.max_column + 1):
        size = sheet.cell(row=1, column=col).value
        if size:
            sizes.append(size)
    return sizes


def parse_inventory(sheet, sizes):
    for row in range(2, sheet.max_row + 1):
        uniform_name = sheet.cell(row=row, column=1).value
        if not uniform_name:
            continue

        uniform = get_or_create_uniform_type(uniform_name)

        for i, size in enumerate(sizes):
            qty = sheet.cell(row=row, column=i + 2).value
            if qty and qty > 0:
                upsert_inventory(uniform.id, size, qty)


def get_or_create_uniform_type(name):
    uniform = UniformType.query.filter_by(name=name).first()
    if uniform:
        return uniform

    uniform = UniformType(name=name)
    db.session.add(uniform)
    return uniform


def upsert_inventory(uniform_type_id, size, quantity):
    record = SupplyInventory.query.filter_by(
        uniform_type_id=uniform_type_id,
        size=size
    ).first()

    if record:
        record.quantity = quantity
    else:
        record = SupplyInventory(
            uniform_type_id=uniform_type_id,
            size=size,
            quantity=quantity
        )
        db.session.add(record)
