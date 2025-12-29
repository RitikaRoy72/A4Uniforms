from openpyxl import load_workbook
from models import db, Cadet, UniformType, IssuedUniform


CADET_COL_COUNT = 5


def import_cadet_sheet(file):
    wb = load_workbook(file)
    sheet = wb.active

    uniform_columns = parse_uniform_headers(sheet)
    parse_cadets(sheet, uniform_columns)

    db.session.commit()


def parse_uniform_headers(sheet):
    uniform_columns = {}

    for col in range(CADET_COL_COUNT + 1, sheet.max_column + 1):
        uniform_name = sheet.cell(row=1, column=col).value
        if uniform_name:
            uniform = get_or_create_uniform_type(uniform_name)
            uniform_columns[col] = uniform.id

    return uniform_columns


def parse_cadets(sheet, uniform_columns):
    for row in range(2, sheet.max_row + 1):
        cadet = get_or_create_cadet(sheet, row)

        for col, uniform_type_id in uniform_columns.items():
            size = sheet.cell(row=row, column=col).value
            if size:
                issue_uniform(cadet.id, uniform_type_id, size)


def get_or_create_cadet(sheet, row):
    cadet_id = str(sheet.cell(row, 1).value).strip()

    cadet = Cadet.query.filter_by(cadet_id=cadet_id).first()
    if cadet:
        return cadet

    cadet = Cadet(
        cadet_id=cadet_id,
        last_name=sheet.cell(row, 2).value,
        first_name=sheet.cell(row, 3).value,
        rank=sheet.cell(row, 4).value,
        flight=sheet.cell(row, 5).value,
    )
    db.session.add(cadet)
    return cadet


def get_or_create_uniform_type(name):
    uniform = UniformType.query.filter_by(name=name).first()
    if uniform:
        return uniform

    uniform = UniformType(name=name)
    db.session.add(uniform)
    return uniform


def issue_uniform(cadet_id, uniform_type_id, size):
    existing = IssuedUniform.query.filter_by(
        cadet_id=cadet_id,
        uniform_type_id=uniform_type_id
    ).first()

    if existing:
        existing.size = size
        return

    issued = IssuedUniform(
        cadet_id=cadet_id,
        uniform_type_id=uniform_type_id,
        size=size
    )
    db.session.add(issued)
