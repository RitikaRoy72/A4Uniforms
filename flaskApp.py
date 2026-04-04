from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify)
from models import db, Cadet, IssuedUniform, SupplyInventory, UniformRequest, UniformType
import os, csv
from datetime import datetime
from excel_parsers import load_cadet_excel, clean_uniform_name
import openpyxl
from flask import send_file
import auth
from inventory_parser import load_inventory
from mailer import init_mail, send_reset_email

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-in-prod")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///uniforms.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)
init_mail(app)

with app.app_context():
    db.create_all()

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "det520admin")

INVENTORY_FILE = "data/AllInventoryData.xlsx"
REQUESTS_FILE  = "data/uniform_requests.csv"
os.makedirs("data", exist_ok=True)


# ── Helpers ──────────────────────────────────────────────────────────────────

def is_admin():
    return session.get("role") == "admin"

def current_cadet_name():
    return session.get("cadet_name")

def all_cadet_names():
    names = set()
    if not os.path.isdir("cadet_data"):
        return names
    for filename in os.listdir("cadet_data"):
        if not filename.endswith(".xlsx"):
            continue
        data = load_cadet_excel(os.path.join("cadet_data", filename))
        for row in data["rows"]:
            n = row["cadet"].get("name")
            if n:
                names.add(n.strip())
    return names

def parse_display_name(raw):
    """'Allen,Michael' -> 'Michael Allen'"""
    if raw and "," in raw:
        last, first = raw.split(",", 1)
        return f"{first.strip()} {last.strip()}"
    return raw or ""

def _resolve_name(typed):
    """Match typed name to an Excel cadet name key. Returns key or None."""
    typed = typed.strip()
    all_names = all_cadet_names()
    typed_lower = typed.lower()
    for n in all_names:
        if n.lower() == typed_lower:
            return n
        if " " in typed:
            parts = typed.split()
            reversed_form = f"{parts[-1]},{' '.join(parts[:-1])}"
            if n.lower() == reversed_form.lower():
                return n
        if "," in n:
            last = n.split(",")[0]
            if last.lower() == typed_lower:
                return n
    return None


# ── Home ─────────────────────────────────────────────────────────────────────

@app.route("/")
def home():
    return redirect(url_for("cadet"))


# ── Upload ───────────────────────────────────────────────────────────────────

UPLOAD_TARGETS = {
    "cadets_blues": "cadet_data/CadetData_Blues.xlsx",
    "cadets_ocp":   "cadet_data/CadetData_OCP.xlsx",
    "cadets_ptg":   "cadet_data/CadetData_PTG.xlsx",
    "inventory":    "data/AllInventoryData.xlsx",
}

@app.route("/upload", methods=["GET", "POST"])
def upload():
    if request.method == "POST":
        action = request.form.get("action")
        target = request.form.get("target")
        if target not in UPLOAD_TARGETS:
            return "Invalid target", 400
        path = UPLOAD_TARGETS[target]
        if action == "download":
            if not os.path.exists(path):
                return "File not found — nothing has been uploaded yet.", 404
            return send_file(path, as_attachment=True)
        if action == "upload":
            file = request.files.get("file")
            if not file or file.filename == "":
                return "No file provided.", 400
            file.save(path)
            return render_template("upload.html", message="Uploaded successfully.")
    return render_template("upload.html")


# ── Cadets: main page ─────────────────────────────────────────────────────────

@app.route("/cadets", methods=["GET", "POST"])
def cadet():
    error = None

    if request.method == "POST":
        action = request.form.get("action")

        if action == "admin_login":
            u = request.form.get("username", "").strip()
            p = request.form.get("password", "").strip()
            if u == ADMIN_USERNAME and p == ADMIN_PASSWORD:
                session["role"] = "admin"
                session.pop("cadet_name", None)
                return redirect(url_for("cadet"))
            error = "Invalid admin credentials."

        elif action == "cadet_login":
            name = request.form.get("cadet_name", "").strip()
            pwd  = request.form.get("cadet_password", "").strip()
            excel_key = _resolve_name(name)
            if not excel_key:
                error = "Cadet not found. Check your name spelling."
            elif not auth.is_registered(excel_key):
                session["pending_registration"] = excel_key
                return redirect(url_for("register_page"))
            elif not auth.verify(excel_key, pwd):
                error = "Incorrect password."
            else:
                session["role"] = "cadet"
                session["cadet_name"] = excel_key
                return redirect(url_for("cadet"))

        elif action == "logout":
            session.clear()
            return redirect(url_for("cadet"))

    role = session.get("role")
    cadet_data = {}
    if os.path.isdir("cadet_data"):
        for filename in os.listdir("cadet_data"):
            if not filename.endswith(".xlsx"):
                continue
            category = filename.replace("CadetData_", "").replace(".xlsx", "")
            filepath = os.path.join("cadet_data", filename)
            try:
                cadet_data[category] = load_cadet_excel(filepath)
            except Exception as e:
                print(f"Warning: could not read {filepath}: {e}")

    logged_in_name = current_cadet_name()
    display_name   = parse_display_name(logged_in_name) if logged_in_name else None

    inventory_data = load_inventory(INVENTORY_FILE) if role == "admin" else []

    pending_requests = []
    if role == "admin" and os.path.exists(REQUESTS_FILE):
        with open(REQUESTS_FILE, newline="", encoding="utf-8") as f:
            pending_requests = list(csv.DictReader(f))

    return render_template(
        "cadets.html",
        cadet_data=cadet_data,
        role=role,
        cadet_name=logged_in_name,
        display_name=display_name,
        inventory_data=inventory_data,
        requests=pending_requests,
        error=error,
    )


# ── First-time registration ───────────────────────────────────────────────────

@app.route("/cadets/register", methods=["GET", "POST"])
def register_page():
    pending = session.get("pending_registration")
    if not pending:
        return redirect(url_for("cadet"))

    error   = None
    display = parse_display_name(pending)

    if request.method == "POST":
        email   = request.form.get("email", "").strip()
        pwd     = request.form.get("password", "").strip()
        confirm = request.form.get("confirm", "").strip()

        if not email or "@" not in email:
            error = "Please enter a valid email address."
        elif len(pwd) < 8:
            error = "Password must be at least 8 characters."
        elif pwd != confirm:
            error = "Passwords do not match."
        else:
            ok = auth.register(pending, email, pwd)
            if ok:
                session.pop("pending_registration", None)
                session["role"] = "cadet"
                session["cadet_name"] = pending
                return redirect(url_for("cadet"))
            else:
                error = "This account is already registered. Please log in."

    return render_template("register.html", display_name=display, error=error)


# ── Forgot password ───────────────────────────────────────────────────────────

@app.route("/cadets/forgot", methods=["GET", "POST"])
def forgot_password():
    message = None
    error   = None

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        key   = auth.find_name_by_email(email)
        if key:
            token     = auth.create_reset_token(key)
            reset_url = url_for("reset_password", token=token, _external=True)
            sent      = send_reset_email(email, parse_display_name(key), reset_url)
            if sent:
                message = "Reset link sent! Check your email."
            else:
                message = (
                    f"Email sending is not configured. "
                    f"Use this link to reset your password (expires in 30 min):<br/>"
                    f"<a href='{reset_url}' style='color:#4a90d9'>{reset_url}</a>"
                )
        else:
            message = "If that email is registered, a reset link has been sent."

    return render_template("forgot.html", message=message, error=error)


# ── Reset password ────────────────────────────────────────────────────────────

@app.route("/cadets/reset/<token>", methods=["GET", "POST"])
def reset_password(token):
    key = auth.validate_reset_token(token)
    if not key:
        return render_template("reset.html", error="This link is invalid or has expired.",
                               expired=True, token=token)
    error = None
    if request.method == "POST":
        pwd     = request.form.get("password", "").strip()
        confirm = request.form.get("confirm", "").strip()
        if len(pwd) < 8:
            error = "Password must be at least 8 characters."
        elif pwd != confirm:
            error = "Passwords do not match."
        else:
            ok = auth.apply_reset(token, pwd)
            if ok:
                return redirect(url_for("cadet"))
            error = "Reset failed — link may have already been used."

    return render_template("reset.html", token=token, error=error,
                           display_name=parse_display_name(key))



# ── Add cadet ─────────────────────────────────────────────────────────────────

import shutil, tempfile

YEAR_ORDER = {"AS100": 1, "AS200": 2, "AS300": 3, "AS400": 4}

def sort_key(row_tuple):
    """Sort by AS year (AS100 first) then alphabetically by name."""
    name, status = row_tuple[0], row_tuple[1]
    year_rank = YEAR_ORDER.get(str(status).strip().upper(), 99) if status else 99
    return (year_rank, str(name).lower() if name else "")

def rewrite_excel_sorted(filepath, new_row=None, delete_name=None, update=None):
    """
    Read an Excel file, optionally add/delete/update a cadet row,
    sort by year then name, and write back to a clean workbook.
    Uses data_only=True + fresh workbook to avoid corrupting files
    that contain comments or drawings (openpyxl cannot round-trip those).
    Backs up AFTER a successful read, so we never back up or overwrite
    a file we could not parse.
    """
    # --- Read first; raise immediately if file is unreadable ---
    wb_read = openpyxl.load_workbook(filepath, data_only=True)
    ws_read = wb_read.active

    header    = [c.value for c in ws_read[1]]
    data_rows = []
    for row in ws_read.iter_rows(min_row=2, values_only=True):
        if not any(row):
            break
        data_rows.append(list(row))
    wb_read.close()

    # Apply delete
    if delete_name:
        data_rows = [r for r in data_rows
                     if str(r[0] or "").strip().lower() != delete_name.strip().lower()]

    # Apply update
    if update:
        for r in data_rows:
            if str(r[0] or "").strip().lower() == update["name"].strip().lower():
                field = update["field"]
                if field == "status":
                    r[1] = update["value"]
                elif field == "gender":
                    r[2] = update["value"]
                break

    # Apply add
    if new_row:
        data_rows.append(new_row)

    # Sort: (year_rank, name)
    data_rows.sort(key=lambda r: sort_key((r[0], r[1])))

    # Write to a temp file first, then atomically replace the original.
    # This prevents Windows file-locking from producing a half-written,
    # corrupted file when Werkzeug or another process has the original open.
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active

    for j, val in enumerate(header):
        ws_new.cell(row=1, column=j + 1).value = val

    for i, row in enumerate(data_rows):
        for j, val in enumerate(row):
            ws_new.cell(row=2 + i, column=j + 1).value = val

    dir_name = os.path.dirname(os.path.abspath(filepath))
    tmp_fd, tmp_path = tempfile.mkstemp(dir=dir_name, suffix=".xlsx")
    try:
        os.close(tmp_fd)
        wb_new.save(tmp_path)
        shutil.move(tmp_path, filepath)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


@app.route("/cadets/add", methods=["POST"])
def add_cadet():
    if not is_admin():
        return jsonify({"ok": False, "error": "Unauthorized"}), 403

    data   = request.get_json()
    name   = data.get("name", "").strip()
    status = data.get("status", "AS100").strip().upper()
    gender = data.get("gender", "").strip().upper()

    if not name:
        return jsonify({"ok": False, "error": "Name is required"}), 400

    # Normalise to Last,First format if space-separated
    if " " in name and "," not in name:
        parts = name.split()
        name = f"{parts[-1]},{' '.join(parts[:-1])}"

    skipped = []
    for filename in os.listdir("cadet_data"):
        if not filename.endswith(".xlsx"):
            continue
        filepath = os.path.join("cadet_data", filename)
        try:
            # Use data_only=True — same mode as rewrite_excel_sorted uses,
            # avoids triggering CRC/zip errors on files with comments/drawings
            wb      = openpyxl.load_workbook(filepath, data_only=True)
            ws      = wb.active
            n_cols  = ws.max_column
            new_row = [name, status, gender] + [""] * (n_cols - 3)
            wb.close()
            rewrite_excel_sorted(filepath, new_row=new_row)
        except Exception as e:
            print(f"Warning: skipping corrupted file {filename}: {e}")
            skipped.append(filename)

    warning = f" (skipped corrupted file(s): {chr(39)}{chr(39).join(skipped)}{chr(39)})" if skipped else ""
    return jsonify({"ok": True, "name": name, "status": status, "gender": gender, "warning": warning})


@app.route("/cadets/delete", methods=["POST"])
def delete_cadet():
    if not is_admin():
        return jsonify({"ok": False, "error": "Unauthorized"}), 403

    data = request.get_json()
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"ok": False, "error": "Name required"}), 400

    for filename in os.listdir("cadet_data"):
        if not filename.endswith(".xlsx"):
            continue
        filepath = os.path.join("cadet_data", filename)
        try:
            rewrite_excel_sorted(filepath, delete_name=name)
        except Exception as e:
            pass  # file may not have this cadet, that is fine

    return jsonify({"ok": True})


@app.route("/cadets/update-field", methods=["POST"])
def update_cadet_field():
    """Update status or gender for a cadet across all sheets, then re-sort."""
    if not is_admin():
        return jsonify({"ok": False, "error": "Unauthorized"}), 403

    data  = request.get_json()
    name  = data.get("name", "").strip()
    field = data.get("field", "")   # "status" or "gender"
    value = data.get("value", "").strip()

    if field not in ("status", "gender"):
        return jsonify({"ok": False, "error": "Invalid field"}), 400

    # Only update the category sheet the save came from, then re-sort all
    category = data.get("category", "")
    filepath = os.path.join("cadet_data", f"CadetData_{category}.xlsx")
    try:
        rewrite_excel_sorted(filepath, update={"name": name, "field": field, "value": value})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True})

# ── Save cadet edits ──────────────────────────────────────────────────────────

@app.route("/cadets/save", methods=["POST"])
def save_cadet_edits():
    role = session.get("role")
    if role not in ("cadet", "admin"):
        return jsonify({"ok": False, "error": "Not logged in"}), 403

    data       = request.get_json()
    cadet_name = data.get("cadet_name", "").strip()
    category   = data.get("category", "")
    items      = data.get("items", [])

    if role == "cadet" and cadet_name.lower() != (current_cadet_name() or "").lower():
        return jsonify({"ok": False, "error": "Unauthorized"}), 403

    filepath = os.path.join("cadet_data", f"CadetData_{category}.xlsx")
    try:
        # Read with data_only=True to avoid CRC/zip errors on files with comments
        wb      = openpyxl.load_workbook(filepath, data_only=True)
        ws      = wb.active
        headers = [c.value.strip() if isinstance(c.value, str) else "" for c in ws[1]]

        # Build a full data snapshot so we can pass it through rewrite_excel_sorted
        all_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                break
            all_rows.append(list(row))
        wb.close()

        # Apply edits to the in-memory snapshot
        for r in all_rows:
            if str(r[0] or "").strip().lower() == cadet_name.lower():
                for edit in items:
                    for idx, h in enumerate(headers):
                        if h and edit["item"].lower() in h.lower():
                            if idx < len(r):
                                r[idx] = edit.get("size", "")
                            if idx + 1 < len(r):
                                r[idx + 1] = edit.get("qty", "")
                            break
                break
        else:
            return jsonify({"ok": False, "error": "Cadet not found in sheet"}), 404

        # Write to temp file then atomically replace — prevents Windows
        # file-locking from truncating the file mid-write.
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        for j, val in enumerate(headers):
            ws_new.cell(row=1, column=j + 1).value = val
        for i, row in enumerate(all_rows):
            for j, val in enumerate(row):
                ws_new.cell(row=2 + i, column=j + 1).value = val
        dir_name = os.path.dirname(os.path.abspath(filepath))
        tmp_fd, tmp_path = tempfile.mkstemp(dir=dir_name, suffix=".xlsx")
        try:
            os.close(tmp_fd)
            wb_new.save(tmp_path)
            shutil.move(tmp_path, filepath)
        except Exception:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            raise

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Submit uniform request ────────────────────────────────────────────────────


@app.route("/cadets/request", methods=["POST"])
def cadet_request():
    role = session.get("role")
    if role not in ("cadet", "admin"):
        return jsonify({"ok": False, "error": "Not logged in"}), 403

    data = request.get_json()
    raw_name = data.get("cadet_name", "")
    # parse "Last,First" into separate fields for consistent CSV columns
    if "," in raw_name:
        last, first = raw_name.split(",", 1)
    else:
        parts = raw_name.split()
        first = parts[0] if parts else ""
        last  = " ".join(parts[1:]) if len(parts) > 1 else ""
    row  = {
        "timestamp":    datetime.now().isoformat(),
        "first_name":   first.strip(),
        "last_name":    last.strip(),
        "rank":         "",
        "uniform_type": "",
        "uniform_item": data.get("uniform_item", ""),
        "size":         data.get("size", ""),
        "reason":       data.get("reason", ""),
        "status":       "PENDING"
    }
    file_exists = os.path.exists(REQUESTS_FILE)
    with open(REQUESTS_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=row.keys())
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)
    return jsonify({"ok": True})


# ── Inventory ─────────────────────────────────────────────────────────────────

@app.route("/inventory")
def inventory():
    inventory_data = load_inventory_excel("data/AllInventoryData.xlsx")
    return render_template("inventory.html", inventory_data=inventory_data)

def load_inventory_excel(filepath):
    wb   = openpyxl.load_workbook(filepath, data_only=True)
    data = {}
    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = []
        empty_streak = 0
        for row in ws.iter_rows(values_only=True):
            if not any(row):
                empty_streak += 1
                if empty_streak >= 4:
                    break
                rows.append({"type": "break"})
                continue
            empty_streak = 0
            if row[1] and not any(row[2:]):
                rows.append({"type": "section", "title": row[1]})
            else:
                rows.append({"type": "data", "cells": list(row)})
        data[sheet_name] = rows
    return data


# ── Requests ──────────────────────────────────────────────────────────────────


@app.route("/cadets/requests/resolve", methods=["POST"])
def resolve_request():
    if not is_admin():
        return jsonify({"ok": False, "error": "Unauthorized"}), 403

    data  = request.get_json()
    index = data.get("index")

    if not os.path.exists(REQUESTS_FILE):
        return jsonify({"ok": False, "error": "No requests file found"}), 404

    with open(REQUESTS_FILE, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []
        rows   = list(reader)

    if index is None or index < 0 or index >= len(rows):
        return jsonify({"ok": False, "error": "Invalid index"}), 400

    rows.pop(index)

    with open(REQUESTS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    return jsonify({"ok": True})

@app.route("/requests", methods=["GET", "POST"])
def requests_page():
    if request.method == "POST":
        os.makedirs(os.path.dirname(REQUESTS_FILE), exist_ok=True)
        row = {
            "timestamp":    datetime.now().isoformat(),
            "first_name":   request.form.get("first_name", ""),
            "last_name":    request.form.get("last_name", ""),
            "rank":         request.form.get("rank", ""),
            "uniform_type": request.form.get("uniform_type", ""),
            "uniform_item": request.form.get("uniform_item", ""),
            "size":         request.form.get("size", ""),
            "reason":       request.form.get("reason", ""),
            "status":       "PENDING"
        }
        file_exists = os.path.exists(REQUESTS_FILE)
        with open(REQUESTS_FILE, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=row.keys())
            if not file_exists:
                writer.writeheader()
            writer.writerow(row)
        return redirect(url_for("requests_page"))
    return render_template("requests.html")


# ── Approve ───────────────────────────────────────────────────────────────────

@app.route("/approve", methods=["GET", "POST"])
def approve_requests():
    CSV_FILE = "data/uniform_requests.csv"
    if request.method == "POST":
        if request.form.get("reset") == "reset":
            with open(CSV_FILE, newline="", encoding="utf-8") as f:
                headers = next(csv.reader(f))
            with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(headers)
            return redirect(url_for("approve_requests"))

    if not os.path.exists(CSV_FILE):
        reqs = []
    else:
        with open(CSV_FILE, newline="", encoding="utf-8") as f:
            reqs = list(csv.DictReader(f))
    return render_template("approve.html", requests=reqs)


if __name__ == "__main__":
    app.run(debug=True)