from openpyxl import load_workbook
from models import db, Cadet, IssuedUniform, UniformType

def import_cadet_sheet(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    uniform_headers = headers[5:]  # uniforms start after cadet info

    # Ensure uniform types exist
    uniform_map = {}
    for name in uniform_headers:
        ut = UniformType.query.filter_by(name=name).first()
        if not ut:
            ut = UniformType(name=name)
            db.session.add(ut)
        uniform_map[name] = ut
    db.session.commit()

    for row in ws.iter_rows(min_row=2, values_only=True):
        cadet = Cadet(
            cadet_id=row[0],
            first_name=row[1],
            last_name=row[2],
            rank=row[3],
            flight=row[4]
        )
        db.session.add(cadet)
        db.session.commit()

        for idx, size in enumerate(row[5:]):
            if size:
                issued = IssuedUniform(
                    cadet_id=cadet.id,
                    uniform_type_id=uniform_map[uniform_headers[idx]].id,
                    size=size
                )
                db.session.add(issued)

    db.session.commit()
