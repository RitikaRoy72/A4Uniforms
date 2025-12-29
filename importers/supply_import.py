from openpyxl import load_workbook
from models import db, UniformType, SupplyInventory

def import_supply_sheet(filepath):
    wb = load_workbook(filepath)
    ws = wb["Supply"]  # name this sheet explicitly

    sizes = [cell.value for cell in ws[1]][1:]

    for row in ws.iter_rows(min_row=2):
        uniform_name = row[0].value

        ut = UniformType.query.filter_by(name=uniform_name).first()
        if not ut:
            ut = UniformType(name=uniform_name)
            db.session.add(ut)
            db.session.commit()

        for i, cell in enumerate(row[1:]):
            if cell.value:
                inv = SupplyInventory(
                    uniform_type_id=ut.id,
                    size=sizes[i],
                    quantity=cell.value
                )
                db.session.add(inv)

    db.session.commit()
