import openpyxl
import re

def clean_uniform_name(header):
    header = re.sub(r"\(.*?\)", "", header)
    header = re.sub(r"\s*-\s*Size.*$", "", header)
    header = re.sub(r"\s*Size.*$", "", header)
    return header.strip()

def load_cadet_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [
        cell.value.strip() if isinstance(cell.value, str) else ""
        for cell in ws[1]
    ]

    uniform_items = []
    column_map = {}

    i = 3  # start after Name / Status / Gender
    while i < len(headers) - 1:
        h = headers[i]
        next_h = headers[i + 1]

        # Detect Size/Qty column pairs by position
        if re.search(r"size", h, re.IGNORECASE) and re.search(r"qty|quantity", next_h, re.IGNORECASE):
            item = clean_uniform_name(h)
            uniform_items.append(item)
            column_map[item] = (i, i + 1)
            i += 2
        else:
            i += 1

    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row[0:3]):
            break

        cadet = {
            "name": row[0],
            "status": row[1],
            "gender": row[2],
        }

        items = {}
        for item, (size_col, qty_col) in column_map.items():
            items[item] = {
                "size": row[size_col] if size_col < len(row) else "",
                "qty": row[qty_col] if qty_col < len(row) else "",
            }

        rows.append({
            "cadet": cadet,
            "items": items
        })

    return {
        "uniform_items": uniform_items,
        "rows": rows
    }
